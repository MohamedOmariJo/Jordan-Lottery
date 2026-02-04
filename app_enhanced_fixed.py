"""
=============================================================================
🎰 تطبيق اليانصيب الأردني المتطور - النسخة 2.0
=============================================================================
تطبيق شامل لتوليد وتحليل أرقام اليانصيب مع تحليلات متقدمة

المطور: محمد العمري
التاريخ: فبراير 2026
الإصدار: 2.0.0
=============================================================================
"""

import streamlit as st
import pandas as pd
import numpy as np
import random
import time
import logging
import io
from datetime import datetime, timedelta
from collections import Counter, defaultdict
from typing import List, Dict, Optional, Tuple, Set, Union
from itertools import chain, combinations

# رسوم بيانية
import plotly.express as px
import plotly.graph_objects as go

# تصدير Excel
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# تصدير PDF
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT
    PDF_AVAILABLE = True
except:
    PDF_AVAILABLE = False

import warnings
warnings.filterwarnings('ignore')

# ==============================================================================
# 1. إعدادات النظام
# ==============================================================================

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(message)s', datefmt='%H:%M:%S')
logger = logging.getLogger("JordanLottery")

class LotteryConfig:
    MIN_NUM = 1
    MAX_NUM = 32
    DEFAULT_TICKET_SIZE = 6
    MIN_TICKET_SIZE = 6
    MAX_TICKET_SIZE = 10
    MAX_GENERATION_ATTEMPTS = 50000
    STRICT_SHADOW_ATTEMPTS = 15000
    DEFAULT_SUM_TOLERANCE = 0.15
    MAX_BATCH_SIZE = 10
    
    TICKET_PRICES = {6: 1, 7: 7, 8: 28, 9: 84, 10: 210}
    MATCH_PRIZES = {3: 1, 4: 15, 5: 500, 6: "JACKPOT"}

def initialize_session_state():
    defaults = {
        'history_df': None,
        'analyzer': None,
        'generator': None,
        'last_result': None,
        'theme': 'light',
        'hot_color': '#22c55e',
        'cold_color': '#3b82f6',
        'balanced_color': '#f59e0b',
        'generated_tickets_session': [],
    }
    for key, value in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = value

# ==============================================================================
# 2. الثيم والتنسيقات
# ==============================================================================

def apply_theme():
    theme = st.session_state.theme
    
    if theme == 'dark':
        bg_color = '#1e1e1e'
        text_color = '#ffffff'
        card_bg = '#2d2d2d'
        border_color = '#404040'
    else:
        bg_color = '#ffffff'
        text_color = '#1f2937'
        card_bg = '#f9fafb'
        border_color = '#e5e7eb'
    
    st.markdown(f"""
    <style>
        .stApp {{
            background-color: {bg_color};
            color: {text_color};
        }}
        
        @keyframes fadeIn {{
            from {{ opacity: 0; transform: translateY(10px); }}
            to {{ opacity: 1; transform: translateY(0); }}
        }}
        
        .number-animated {{
            animation: fadeIn 0.4s ease-out;
            display: inline-block;
            margin: 3px;
        }}
        
        .lottery-number {{
            display: inline-block;
            background: {st.session_state.hot_color};
            color: white;
            padding: 8px 14px;
            margin: 3px;
            border-radius: 50%;
            font-weight: bold;
            border: 2px solid rgba(255,255,255,0.3);
            box-shadow: 0 2px 5px rgba(0,0,0,0.2);
        }}
        
        .footer {{
            position: fixed;
            bottom: 0;
            left: 0;
            width: 100%;
            background-color: {card_bg};
            color: {text_color};
            text-align: center;
            padding: 10px;
            font-size: 14px;
            border-top: 1px solid {border_color};
            z-index: 999;
        }}
    </style>
    """, unsafe_allow_html=True)

# ==============================================================================
# 3. تحميل البيانات
# ==============================================================================

@st.cache_data(show_spinner=False)
def load_and_process_data(file_input):
    try:
        is_csv = False
        if isinstance(file_input, str):
            is_csv = file_input.endswith('.csv')
        else:
            is_csv = file_input.name.endswith('.csv')

        if is_csv:
            df = pd.read_csv(file_input)
        else:
            df = pd.read_excel(file_input)
        
        df.dropna(how='all', inplace=True)
        
        required_cols = ['N1', 'N2', 'N3', 'N4', 'N5', 'N6']
        if not set(required_cols).issubset(df.columns):
            return None, "خطأ: الملف لا يحتوي على أعمدة الأرقام (N1...N6)"

        for col in required_cols:
            df[col] = pd.to_numeric(df[col], errors='coerce')
        
        df.dropna(subset=required_cols, inplace=True)
        df['numbers'] = df[required_cols].values.tolist()
        
        def is_valid_draw(nums):
            return all(LotteryConfig.MIN_NUM <= int(n) <= LotteryConfig.MAX_NUM for n in nums)

        df = df[df['numbers'].apply(is_valid_draw)]
        
        if df.empty:
            return None, "خطأ: لا توجد بيانات صالحة"

        df['numbers'] = df['numbers'].apply(lambda x: sorted([int(n) for n in x]))
        
        if 'رقم السحب' in df.columns:
            df = df.rename(columns={'رقم السحب': 'draw_id'})
        elif 'DrawID' in df.columns:
            df = df.rename(columns={'DrawID': 'draw_id'})
        elif 'draw_id' not in df.columns:
            df['draw_id'] = range(1, len(df) + 1)
        
        if 'تاريخ السحب' in df.columns:
            df['date'] = pd.to_datetime(df['تاريخ السحب'], errors='coerce')
        elif 'date' not in df.columns:
            start_date = datetime(2023, 9, 17)
            df['date'] = [start_date + timedelta(days=i*3) for i in range(len(df))]
        
        df['sum'] = df['numbers'].apply(sum)
        df['odd_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 1))
        df['even_count'] = df['numbers'].apply(lambda x: sum(1 for n in x if n % 2 == 0))
        
        return df, "Success"
        
    except Exception as e:
        logger.error(f"Error: {e}")
        return None, f"خطأ: {str(e)}"

# ==============================================================================
# 4. المحلل المتقدم
# ==============================================================================

class AdvancedAnalyzer:
    def __init__(self, history_df: pd.DataFrame):
        self.history_df = history_df
        self.past_draws_sets = [set(nums) for nums in history_df['numbers']]
        self.draw_map = {row['draw_id']: row['numbers'] for _, row in history_df.iterrows()}
        
        self.number_to_draws_index = defaultdict(set)
        for idx, draw_set in enumerate(self.past_draws_sets):
            for num in draw_set:
                self.number_to_draws_index[num].add(idx)
        
        all_numbers = list(chain.from_iterable(history_df['numbers']))
        self.frequency = Counter(all_numbers)
        self.total_draws = len(history_df)
        
        all_sums = [sum(nums) for nums in history_df['numbers']]
        self.global_avg_sum = sum(all_sums) / len(all_sums) if all_sums else 0
        
        sorted_nums = sorted(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1), 
                           key=lambda x: self.frequency[x], reverse=True)
        self.hot_pool = set(sorted_nums[:16])
        self.cold_pool = set(sorted_nums[16:])
        
        self._calculate_gaps()
        self._analyze_combinations()
    
    def _calculate_gaps(self):
        self.gaps = {}
        for num in range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1):
            last_appearance = -1
            for idx in range(self.total_draws - 1, -1, -1):
                if num in self.history_df.iloc[idx]['numbers']:
                    last_appearance = idx
                    break
            
            if last_appearance == -1:
                self.gaps[num] = self.total_draws
            else:
                self.gaps[num] = self.total_draws - 1 - last_appearance
    
    def _analyze_combinations(self):
        self.pair_frequency = Counter()
        for nums in self.history_df['numbers']:
            for pair in combinations(nums, 2):
                self.pair_frequency[pair] += 1
        
        self.triplet_frequency = Counter()
        for nums in self.history_df['numbers']:
            for triplet in combinations(nums, 3):
                self.triplet_frequency[triplet] += 1
    
    def get_ticket_profile(self, ticket: List[int]) -> str:
        hot_count = sum(1 for n in ticket if n in self.hot_pool)
        total = len(ticket)
        if hot_count >= total * 0.7:
            return "🔥 ساخنة"
        elif hot_count <= total * 0.3:
            return "❄️ باردة"
        else:
            return "⚖️ متوازنة"
    
    def calculate_custom_average(self, mode: str, param1: int = 0, param2: int = 0):
        df = self.history_df.copy()
        if mode == "Last N Draws":
            if param1 > len(df):
                param1 = len(df)
            df = df.iloc[-param1:]
        elif mode == "Specific Range":
            df = df[(df['draw_id'] >= param1) & (df['draw_id'] <= param2)]
        
        if df.empty:
            return self.global_avg_sum, []
        
        sums = [sum(nums) for nums in df['numbers']]
        avg = sum(sums) / len(sums) if sums else 0
        return avg, sums
    
    def get_numbers_from_draw(self, draw_id: int):
        return self.draw_map.get(draw_id)
    
    def check_matches_history(self, ticket_numbers: List[int]):
        matches_found = {6: [], 5: [], 4: [], 3: []}
        ticket_set = set(ticket_numbers)
        
        for draw_id, draw_nums in self.draw_map.items():
            intersection = ticket_set & set(draw_nums)
            count = len(intersection)
            if count in matches_found:
                matches_found[count].append({
                    'draw_id': draw_id,
                    'matched_nums': sorted(list(intersection))
                })
        
        return matches_found
    
    def get_numbers_frequency_stats(self, ticket_numbers: List[int]):
        stats = []
        for num in ticket_numbers:
            count = self.frequency.get(num, 0)
            gap = self.gaps.get(num, 0)
            stats.append({
                'الرقم': num,
                'عدد مرات الظهور': count,
                'آخر ظهور': gap,
                'التصنيف': 'ساخن' if num in self.hot_pool else 'بارد'
            })
        return pd.DataFrame(stats)
    
    def analyze_sequences_history(self, ticket_numbers: List[int]):
        sorted_nums = sorted(ticket_numbers)
        sequences = {}
        
        i = 0
        while i < len(sorted_nums):
            seq_start = i
            while i < len(sorted_nums) - 1 and sorted_nums[i+1] == sorted_nums[i] + 1:
                i += 1
            
            if i > seq_start:
                seq = tuple(sorted_nums[seq_start:i+1])
                
                full_count = 0
                full_draws = []
                for draw_id, draw_nums in self.draw_map.items():
                    if set(seq).issubset(set(draw_nums)):
                        full_count += 1
                        full_draws.append(draw_id)
                
                sub_pairs = {}
                for j in range(len(seq) - 1):
                    pair = (seq[j], seq[j+1])
                    pair_count = 0
                    pair_draws = []
                    for draw_id, draw_nums in self.draw_map.items():
                        if pair[0] in draw_nums and pair[1] in draw_nums:
                            pair_count += 1
                            pair_draws.append(draw_id)
                    sub_pairs[pair] = {'count': pair_count, 'draws': pair_draws[:5]}
                
                sequences[seq] = {
                    'full_count': full_count,
                    'full_draws': full_draws[:5],
                    'sub': sub_pairs
                }
            i += 1
        
        return sequences
    
    def get_temporal_analysis(self, period: str = 'month'):
        df = self.history_df.copy()
        
        if period == 'month':
            df['period'] = df['date'].dt.to_period('M').astype(str)
        elif period == 'year':
            df['period'] = df['date'].dt.year
        else:
            df['period'] = df['date'].dt.to_period('W').astype(str)
        
        expanded = []
        for _, row in df.iterrows():
            for num in row['numbers']:
                expanded.append({'period': row['period'], 'number': num})
        
        expanded_df = pd.DataFrame(expanded)
        pivot = expanded_df.groupby(['period', 'number']).size().reset_index(name='frequency')
        
        return pivot
    
    def get_odd_even_analysis(self):
        return {
            'avg_odd': self.history_df['odd_count'].mean(),
            'avg_even': self.history_df['even_count'].mean(),
            'distribution': self.history_df['odd_count'].value_counts().sort_index().to_dict()
        }
    
    def calculate_standard_deviation(self):
        number_std = {}
        for num in range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1):
            appearances = [1 if num in draw else 0 for draw in self.history_df['numbers']]
            number_std[num] = np.std(appearances)
        
        sums = [sum(nums) for nums in self.history_df['numbers']]
        sum_std = np.std(sums)
        
        return {
            'number_std': number_std,
            'sum_std': sum_std,
            'sum_mean': np.mean(sums)
        }
    
    def get_top_pairs(self, top_n: int = 20):
        return self.pair_frequency.most_common(top_n)
    
    def get_top_triplets(self, top_n: int = 10):
        return self.triplet_frequency.most_common(top_n)

# ==============================================================================
# 5. مولد الأرقام
# ==============================================================================

class SmartGenerator:
    def __init__(self, analyzer: AdvancedAnalyzer):
        self.analyzer = analyzer
        self.validation_errors = []
    
    def validate_criteria(self, criteria: Dict) -> bool:
        self.validation_errors = []
        size = criteria.get('size', 6)
        
        if size < LotteryConfig.MIN_TICKET_SIZE or size > LotteryConfig.MAX_TICKET_SIZE:
            self.validation_errors.append(f"حجم التذكرة يجب أن يكون بين {LotteryConfig.MIN_TICKET_SIZE} و {LotteryConfig.MAX_TICKET_SIZE}")
        
        odd_count = criteria.get('odd_count', 0)
        if odd_count > size:
            self.validation_errors.append("عدد الفردي أكبر من حجم التذكرة")
        
        seq_count = criteria.get('sequences_count', 0)
        if seq_count >= size:
            self.validation_errors.append("عدد المتتاليات كبير جداً")
        
        shadows = criteria.get('shadows_count', 0)
        if shadows > 3:
            self.validation_errors.append("عدد الظلال لا يمكن أن يتجاوز 3")
        
        inc_draw = criteria.get('include_from_draw')
        inc_count = criteria.get('include_count', 0)
        if inc_draw and inc_count > 0:
            if inc_count > size:
                self.validation_errors.append("عدد الأرقام المثبتة أكبر من حجم التذكرة")
            past_nums = self.analyzer.get_numbers_from_draw(inc_draw)
            if not past_nums:
                self.validation_errors.append(f"السحب رقم {inc_draw} غير موجود")
        
        return len(self.validation_errors) == 0
    
    def generate_ticket(self, criteria: Dict):
        size = criteria['size']
        odd_count = criteria.get('odd_count', size // 2)
        strategy = criteria.get('strategy', 'Any')
        target_avg = criteria.get('target_average', self.analyzer.global_avg_sum)
        sum_near_avg = criteria.get('sum_near_avg', True)
        tolerance = LotteryConfig.DEFAULT_SUM_TOLERANCE
        
        if strategy == 'Hot':
            candidate_pool = list(self.analyzer.hot_pool)
        elif strategy == 'Cold':
            candidate_pool = list(self.analyzer.cold_pool)
        elif strategy == 'Balanced':
            hot_needed = size // 2
            cold_needed = size - hot_needed
            candidate_pool = random.sample(list(self.analyzer.hot_pool), min(hot_needed, len(self.analyzer.hot_pool)))
            candidate_pool += random.sample(list(self.analyzer.cold_pool), min(cold_needed, len(self.analyzer.cold_pool)))
        else:
            candidate_pool = list(range(LotteryConfig.MIN_NUM, LotteryConfig.MAX_NUM + 1))
        
        inc_draw = criteria.get('include_from_draw')
        inc_count = criteria.get('include_count', 0)
        fixed_numbers = []
        
        if inc_draw and inc_count > 0:
            past_nums = self.analyzer.get_numbers_from_draw(inc_draw)
            if past_nums:
                fixed_numbers = random.sample(past_nums, min(inc_count, len(past_nums)))
                candidate_pool = [n for n in candidate_pool if n not in fixed_numbers]
        
        for _ in range(LotteryConfig.MAX_GENERATION_ATTEMPTS):
            ticket = fixed_numbers + random.sample(candidate_pool, size - len(fixed_numbers))
            ticket = sorted(ticket)
            
            ticket_sum = sum(ticket)
            ticket_odd = sum(1 for n in ticket if n % 2 == 1)
            
            if sum_near_avg:
                lower_bound = target_avg * (1 - tolerance)
                upper_bound = target_avg * (1 + tolerance)
                if not (lower_bound <= ticket_sum <= upper_bound):
                    continue
            
            if ticket_odd != odd_count:
                continue
            
            seq_count = self._count_sequences(ticket)
            if seq_count != criteria.get('sequences_count', 0):
                continue
            
            shadows = self._count_shadows(ticket)
            if shadows != criteria.get('shadows_count', 0):
                continue
            
            anti_limit = criteria.get('anti_match_limit', 5)
            if self._check_anti_match(ticket, anti_limit):
                continue
            
            return ticket
        
        return None
    
    def _count_sequences(self, ticket: List[int]) -> int:
        count = 0
        for i in range(len(ticket) - 1):
            if ticket[i + 1] == ticket[i] + 1:
                count += 1
        return count
    
    def _count_shadows(self, ticket: List[int]) -> int:
        shadows = 0
        for i in range(len(ticket)):
            for j in range(i + 1, len(ticket)):
                if abs(ticket[i] - ticket[j]) == 10:
                    shadows += 1
        return shadows
    
    def _check_anti_match(self, ticket: List[int], limit: int) -> bool:
        ticket_set = set(ticket)
        for past_draw in self.analyzer.past_draws_sets:
            match_count = len(ticket_set & past_draw)
            if match_count >= limit:
                return True
        return False
    
    def generate_batch(self, criteria: Dict, count: int):
        if not self.validate_criteria(criteria):
            return {
                'status': 'validation_error',
                'errors': self.validation_errors,
                'tickets': [],
                'generated': 0
            }
        
        tickets = []
        generated_sets = set()
        
        for i in range(count):
            ticket = self.generate_ticket(criteria)
            if ticket:
                ticket_tuple = tuple(ticket)
                if ticket_tuple not in generated_sets:
                    generated_sets.add(ticket_tuple)
                    
                    analysis = {
                        'sum': sum(ticket),
                        'profile': self.analyzer.get_ticket_profile(ticket),
                        'sequences': self._count_sequences(ticket),
                        'shadows': self._count_shadows(ticket),
                        'odd_count': sum(1 for n in ticket if n % 2 == 1),
                        'even_count': sum(1 for n in ticket if n % 2 == 0)
                    }
                    
                    tickets.append({
                        'id': i + 1,
                        'numbers': ticket,
                        'analysis': analysis
                    })
        
        status = 'success' if len(tickets) == count else 'partial_success' if len(tickets) > 0 else 'failed'
        
        return {
            'status': status,
            'tickets': tickets,
            'generated': len(tickets),
            'errors': [] if status == 'success' else ['لم يتم توليد جميع التذاكر']
        }
    
    def estimate_success_probability(self, criteria: Dict):
        if not self.validate_criteria(criteria):
            return {'probability': 0, 'advice': 'معايير غير صالحة'}
        
        success = 0
        for _ in range(100):
            ticket = self.generate_ticket(criteria)
            if ticket:
                success += 1
        
        probability = success
        
        if probability > 70:
            advice = "ممتازة ✅"
        elif probability > 30:
            advice = "جيدة ⚠️"
        else:
            advice = "ضعيفة ❌"
        
        return {'probability': probability, 'advice': advice}

# ==============================================================================
# 6. حاسبة الاحتمالات
# ==============================================================================

class ProbabilityCalculator:
    @staticmethod
    def combination(n: int, r: int) -> int:
        if r > n or r < 0:
            return 0
        if r == 0 or r == n:
            return 1
        
        r = min(r, n - r)
        result = 1
        for i in range(r):
            result = result * (n - i) // (i + 1)
        return result
    
    @staticmethod
    def calculate_match_probability(ticket_size: int, match_count: int) -> float:
        total_numbers = 32
        draw_size = 6
        
        ways_to_match = ProbabilityCalculator.combination(draw_size, match_count)
        ways_to_not_match = ProbabilityCalculator.combination(
            total_numbers - draw_size, 
            ticket_size - match_count
        )
        total_ways = ProbabilityCalculator.combination(total_numbers, ticket_size)
        
        if total_ways == 0:
            return 0.0
        
        prob = (ways_to_match * ways_to_not_match) / total_ways
        return prob
    
    @staticmethod
    def calculate_expected_value(ticket_size: int, jackpot: float = 1000000):
        ticket_cost = LotteryConfig.TICKET_PRICES.get(ticket_size, 0)
        
        expected_value = 0.0
        breakdown = {}
        
        for match in [3, 4, 5, 6]:
            prob = ProbabilityCalculator.calculate_match_probability(ticket_size, match)
            
            if match == 6:
                prize = jackpot
            else:
                prize = LotteryConfig.MATCH_PRIZES[match]
            
            expected_value += prob * prize
            breakdown[f'{match} أرقام'] = {
                'probability': prob,
                'prize': prize,
                'contribution': prob * prize
            }
        
        net_expected = expected_value - ticket_cost
        
        return {
            'ticket_size': ticket_size,
            'ticket_cost': ticket_cost,
            'expected_value': expected_value,
            'net_expected_value': net_expected,
            'breakdown': breakdown,
            'roi': (net_expected / ticket_cost * 100) if ticket_cost > 0 else 0
        }

class DrawSimulator:
    def __init__(self, analyzer: AdvancedAnalyzer):
        self.analyzer = analyzer
    
    def simulate_draws(self, num_simulations: int, ticket: List[int], progress_callback=None):
        matches = {3: 0, 4: 0, 5: 0, 6: 0}
        ticket_set = set(ticket)
        
        for i in range(num_simulations):
            draw = random.sample(range(1, 33), 6)
            match_count = len(ticket_set & set(draw))
            
            if match_count in matches:
                matches[match_count] += 1
            
            if progress_callback and (i + 1) % 1000 == 0:
                progress_callback(i + 1, num_simulations)
        
        results = {}
        for match_level, count in matches.items():
            percentage = (count / num_simulations) * 100
            theoretical_prob = ProbabilityCalculator.calculate_match_probability(len(ticket), match_level) * 100
            
            results[match_level] = {
                'count': count,
                'percentage': percentage,
                'theoretical': theoretical_prob,
                'difference': percentage - theoretical_prob
            }
        
        return {
            'total_simulations': num_simulations,
            'ticket': ticket,
            'results': results
        }

# ==============================================================================
# 7. التصدير
# ==============================================================================

class ReportExporter:
    @staticmethod
    def export_to_excel(tickets: List[Dict], analyzer: AdvancedAnalyzer, filename: str = 'tickets.xlsx'):
        wb = Workbook()
        
        # صفحة المعلومات
        ws_info = wb.active
        ws_info.title = "معلومات"
        
        ws_info['A1'] = 'تقرير تذاكر اليانصيب'
        ws_info['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws_info['A1'].fill = PatternFill(start_color='1F4788', end_color='1F4788', fill_type='solid')
        ws_info['A1'].alignment = Alignment(horizontal='center')
        ws_info.merge_cells('A1:D1')
        
        info_data = [
            ['تاريخ:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['عدد التذاكر:', len(tickets)],
            ['السحوبات التاريخية:', analyzer.total_draws],
            ['المتوسط:', f"{analyzer.global_avg_sum:.2f}"],
        ]
        
        for idx, (label, value) in enumerate(info_data, start=3):
            ws_info[f'A{idx}'] = label
            ws_info[f'B{idx}'] = value
            ws_info[f'A{idx}'].font = Font(bold=True)
        
        # صفحة التذاكر
        ws_tickets = wb.create_sheet("التذاكر")
        
        headers = ['#', 'الأرقام', 'المجموع', 'النوع', 'فردي', 'زوجي', 'متتاليات', 'ظلال']
        ws_tickets.append(headers)
        
        for col_num, header in enumerate(headers, start=1):
            cell = ws_tickets.cell(1, col_num)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        for ticket in tickets:
            nums_str = ', '.join(map(str, ticket['numbers']))
            row_data = [
                ticket['id'],
                nums_str,
                ticket['analysis']['sum'],
                ticket['analysis']['profile'],
                ticket['analysis']['odd_count'],
                ticket['analysis']['even_count'],
                ticket['analysis']['sequences'],
                ticket['analysis']['shadows']
            ]
            ws_tickets.append(row_data)
        
        ws_tickets.column_dimensions['A'].width = 10
        ws_tickets.column_dimensions['B'].width = 30
        for col in ['C', 'D', 'E', 'F', 'G', 'H']:
            ws_tickets.column_dimensions[col].width = 12
        
        # استخدام مجلد مؤقت بدلاً من مسار ثابت
        import tempfile
        output_dir = tempfile.gettempdir()
        output_path = os.path.join(output_dir, filename)
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def export_to_pdf(tickets: List[Dict], analyzer: AdvancedAnalyzer, filename: str = 'tickets.pdf'):
        if not PDF_AVAILABLE:
            return None
        
        # استخدام مجلد مؤقت بدلاً من مسار ثابت
        import tempfile
        output_dir = tempfile.gettempdir()
        output_path = os.path.join(output_dir, filename)
        doc = SimpleDocTemplate(output_path, pagesize=A4, 
                               rightMargin=30, leftMargin=30, 
                               topMargin=50, bottomMargin=30)
        
        styles = getSampleStyleSheet()
        
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            textColor=colors.HexColor('#1F4788'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#4472C4'),
            spaceAfter=12,
            alignment=TA_RIGHT
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=11,
            alignment=TA_RIGHT
        )
        
        story = []
        
        story.append(Paragraph("تقرير تذاكر اليانصيب", title_style))
        story.append(Spacer(1, 12))
        
        info_text = f"""
        <b>التاريخ:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>
        <b>عدد التذاكر:</b> {len(tickets)}<br/>
        <b>السحوبات:</b> {analyzer.total_draws}<br/>
        <b>المتوسط:</b> {analyzer.global_avg_sum:.2f}
        """
        story.append(Paragraph(info_text, normal_style))
        story.append(Spacer(1, 20))
        
        story.append(Paragraph("التذاكر:", heading_style))
        story.append(Spacer(1, 12))
        
        for ticket in tickets:
            data = [
                ['#', f"{ticket['id']}"],
                ['الأرقام', ', '.join(map(str, ticket['numbers']))],
                ['المجموع', str(ticket['analysis']['sum'])],
                ['النوع', ticket['analysis']['profile']],
                ['فردي/زوجي', f"{ticket['analysis']['odd_count']}/{ticket['analysis']['even_count']}"],
                ['متتاليات', str(ticket['analysis']['sequences'])],
                ['ظلال', str(ticket['analysis']['shadows'])]
            ]
            
            table = Table(data, colWidths=[2*inch, 3*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#E7E6E6')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
                ('TOPPADDING', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('BOX', (0, 0), (-1, -1), 2, colors.HexColor('#4472C4')),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 15))
        
        footer_text = f"<i>برمجة: محمد العمري</i>"
        story.append(Spacer(1, 30))
        story.append(Paragraph(footer_text, ParagraphStyle('Footer', fontSize=9, alignment=TA_CENTER)))
        
        doc.build(story)
        return output_path

# يتبع في الجزء الثاني...

# ==============================================================================
# 8. واجهة المستخدم - الدوال الرئيسية
# ==============================================================================

def render_dashboard(analyzer: AdvancedAnalyzer):
    st.header("📊 Dashboard - نظرة عامة")
    
    col1, col2, col3, col4 = st.columns(4)
    
    with col1:
        st.metric("السحوبات", analyzer.total_draws)
    
    with col2:
        st.metric("المتوسط", f"{analyzer.global_avg_sum:.2f}")
    
    with col3:
        odd_even = analyzer.get_odd_even_analysis()
        st.metric("متوسط الفردي", f"{odd_even['avg_odd']:.1f}")
    
    with col4:
        st.metric("متوسط الزوجي", f"{odd_even['avg_even']:.1f}")
    
    st.divider()
    
    col_left, col_right = st.columns([1, 1])
    
    with col_left:
        st.subheader("🔥 أكثر 10 أرقام")
        top_10 = analyzer.frequency.most_common(10)
        df_hot = pd.DataFrame(top_10, columns=['الرقم', 'التكرار'])
        fig_hot = px.bar(df_hot, x='الرقم', y='التكرار', color='التكرار',
                        color_continuous_scale='Reds', text='التكرار')
        fig_hot.update_traces(textposition='outside')
        fig_hot.update_layout(showlegend=False, height=350)
        st.plotly_chart(fig_hot, use_container_width=True)
    
    with col_right:
        st.subheader("❄️ أقل 10 أرقام")
        bottom_10 = analyzer.frequency.most_common()[:-11:-1]
        df_cold = pd.DataFrame(bottom_10, columns=['الرقم', 'التكرار'])
        fig_cold = px.bar(df_cold, x='الرقم', y='التكرار', color='التكرار',
                         color_continuous_scale='Blues', text='التكرار')
        fig_cold.update_traces(textposition='outside')
        fig_cold.update_layout(showlegend=False, height=350)
        st.plotly_chart(fig_cold, use_container_width=True)
    
    st.divider()
    
    st.subheader("🗺️ الخريطة الحرارية (1-32)")
    
    heatmap_data = []
    for row in range(4):
        row_data = []
        for col in range(8):
            num = row * 8 + col + 1
            freq = analyzer.frequency.get(num, 0)
            row_data.append(freq)
        heatmap_data.append(row_data)
    
    fig_heatmap = go.Figure(data=go.Heatmap(
        z=heatmap_data,
        x=[str(i) for i in range(1, 9)],
        y=['1-8', '9-16', '17-24', '25-32'],
        text=[[f"{row*8 + col + 1}" for col in range(8)] for row in range(4)],
        texttemplate="%{text}<br>%{z}",
        textfont={"size": 11},
        colorscale='YlOrRd',
        colorbar=dict(title="التكرار")
    ))
    
    fig_heatmap.update_layout(height=350)
    st.plotly_chart(fig_heatmap, use_container_width=True)
    
    st.divider()
    
    st.subheader("🎲 آخر 5 سحوبات")
    last_5 = analyzer.history_df.tail(5)[['draw_id', 'numbers', 'date', 'sum']].copy()
    last_5['الأرقام'] = last_5['numbers'].apply(lambda x: ', '.join(map(str, x)))
    last_5 = last_5.rename(columns={
        'draw_id': 'السحب',
        'date': 'التاريخ',
        'sum': 'المجموع'
    })
    st.dataframe(last_5[['السحب', 'التاريخ', 'الأرقام', 'المجموع']], 
                hide_index=True, use_container_width=True)


def render_smart_generator(analyzer: AdvancedAnalyzer, generator: SmartGenerator):
    st.header("🎰 المولد الذكي")
    
    col_settings, col_results = st.columns([1, 1.5])
    
    with col_settings:
        st.subheader("⚙️ الإعدادات")
        
        strategy = st.selectbox(
            "🎯 الاستراتيجية",
            ["Any (الكل)", "Hot (ساخنة)", "Cold (باردة)", "Balanced (متوازنة)"]
        )
        strategy_map = {
            "Any (الكل)": "Any",
            "Hot (ساخنة)": "Hot",
            "Cold (باردة)": "Cold",
            "Balanced (متوازنة)": "Balanced"
        }
        
        with st.container(border=True):
            st.markdown("**📊 المتوسط**")
            avg_chk = st.checkbox("الالتزام بالمتوسط", value=True)
            target_avg_val = analyzer.global_avg_sum
            chart_data = []
            
            if avg_chk:
                avg_mode = st.selectbox(
                    "المرجع:",
                    ["كافة السحوبات", "آخر N سحب", "نطاق محدد"]
                )
                
                if avg_mode == "آخر N سحب":
                    n_draws = st.number_input("عدد السحوبات", 5, analyzer.total_draws, 20)
                    target_avg_val, chart_data = analyzer.calculate_custom_average("Last N Draws", param1=n_draws)
                    st.caption(f"المتوسط: **{target_avg_val:.2f}**")
                
                elif avg_mode == "نطاق محدد":
                    c1, c2 = st.columns(2)
                    start_d = c1.number_input("من", 1, analyzer.total_draws, max(1, analyzer.total_draws - 50))
                    end_d = c2.number_input("إلى", 1, analyzer.total_draws, analyzer.total_draws)
                    target_avg_val, chart_data = analyzer.calculate_custom_average("Specific Range", param1=start_d, param2=end_d)
                    st.caption(f"المتوسط: **{target_avg_val:.2f}**")
                
                else:
                    target_avg_val, chart_data = analyzer.calculate_custom_average("All")
                    st.caption(f"المتوسط: **{target_avg_val:.2f}**")
                
                if chart_data:
                    st.line_chart(chart_data, height=120)
        
        with st.container(border=True):
            t_count = st.number_input("عدد التذاكر", 1, 10, 3)
            t_size = st.slider("حجم التذكرة", 6, 10, 6)
            odd = st.number_input("عدد الفردي", 0, t_size, t_size // 2)
            seq = st.number_input("المتتاليات", 0, t_size - 1, 0)
            sha = st.number_input("الظلال", 0, 3, 1)
        
        with st.container(border=True):
            st.markdown("**🔄 تثبيت أرقام**")
            use_past = st.checkbox("من سحب سابق")
            inc_draw = None
            inc_cnt = 0
            
            if use_past:
                c1, c2 = st.columns(2)
                inc_draw = c1.number_input("السحب", 1, analyzer.total_draws, analyzer.total_draws)
                inc_cnt = c2.number_input("العدد", 1, min(6, t_size), 1)
                past_nums = analyzer.get_numbers_from_draw(inc_draw)
                if past_nums:
                    st.caption(f"الأرقام: {past_nums}")
        
        anti = st.slider("تجنب تكرار", 3, t_size, 5)
        
        criteria = {
            'size': t_size,
            'sequences_count': seq,
            'odd_count': odd,
            'shadows_count': sha,
            'anti_match_limit': anti,
            'sum_near_avg': avg_chk,
            'target_average': target_avg_val,
            'include_from_draw': inc_draw if use_past else None,
            'include_count': inc_cnt if use_past else 0,
            'strategy': strategy_map[strategy]
        }
        
        if st.button("🔍 فحص الجدوى"):
            with st.spinner("جاري..."):
                est = generator.estimate_success_probability(criteria)
                color = "green" if est['probability'] > 5 else "red"
                st.markdown(f"**النسبة:** :{color}[{est['probability']}%] ({est['advice']})")
        
        if st.button("🚀 توليد", type="primary", use_container_width=True):
            with st.spinner("جاري..."):
                result = generator.generate_batch(criteria, t_count)
                st.session_state.last_result = result
                
                if result['status'] in ['success', 'partial_success']:
                    st.session_state.generated_tickets_session.extend(result['tickets'])
    
    with col_results:
        st.subheader("📋 النتائج")
        
        if st.session_state.last_result:
            res = st.session_state.last_result
            
            if res['status'] == 'validation_error':
                st.error("❌ خطأ:")
                for err in res['errors']:
                    st.write(f"- {err}")
            
            elif res['status'] == 'failed':
                st.error("❌ فشل")
            
            else:
                if res['status'] == 'partial_success':
                    st.warning(f"⚠️ تم توليد {res['generated']} فقط")
                else:
                    st.success(f"✅ تم توليد {res['generated']} تذاكر!")
                
                for ticket in res['tickets']:
                    with st.expander(f"🎫 تذكرة #{ticket['id']} - {ticket['analysis']['profile']}", expanded=True):
                        profile = ticket['analysis']['profile']
                        if '🔥' in profile:
                            color = st.session_state.hot_color
                        elif '❄️' in profile:
                            color = st.session_state.cold_color
                        else:
                            color = st.session_state.balanced_color
                        
                        numbers_html = "".join([
                            f"<span class='lottery-number number-animated' style='background:{color}'>{n}</span>"
                            for n in ticket['numbers']
                        ])
                        st.markdown(numbers_html, unsafe_allow_html=True)
                        
                        ca, cb, cc = st.columns(3)
                        ca.caption(f"المجموع: {ticket['analysis']['sum']}")
                        cb.caption(f"متتاليات: {ticket['analysis']['sequences']}")
                        cc.caption(f"ظلال: {ticket['analysis']['shadows']}")
        
        if st.session_state.generated_tickets_session:
            st.divider()
            st.markdown(f"**💾 تذاكر الجلسة: {len(st.session_state.generated_tickets_session)}**")
            
            col_exp1, col_exp2, col_exp3 = st.columns(3)
            
            with col_exp1:
                if st.button("📥 Excel", use_container_width=True):
                    with st.spinner("جاري..."):
                        filepath = ReportExporter.export_to_excel(
                            st.session_state.generated_tickets_session,
                            analyzer,
                            f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                        )
                        st.success("✅")
                        with open(filepath, 'rb') as f:
                            st.download_button(
                                "⬇️ تحميل",
                                f,
                                file_name=filepath.split('/')[-1],
                                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                            )
            
            with col_exp2:
                if PDF_AVAILABLE and st.button("📄 PDF", use_container_width=True):
                    with st.spinner("جاري..."):
                        filepath = ReportExporter.export_to_pdf(
                            st.session_state.generated_tickets_session,
                            analyzer,
                            f'tickets_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
                        )
                        if filepath:
                            st.success("✅")
                            with open(filepath, 'rb') as f:
                                st.download_button(
                                    "⬇️ تحميل",
                                    f,
                                    file_name=filepath.split('/')[-1],
                                    mime='application/pdf'
                                )
            
            with col_exp3:
                if st.button("🗑️ مسح", use_container_width=True):
                    st.session_state.generated_tickets_session = []
                    st.rerun()


def render_checker(analyzer: AdvancedAnalyzer):
    st.header("🔍 الفاحص التاريخي")
    
    c_check1, c_check2 = st.columns([1, 2])
    
    with c_check1:
        chk_size = st.radio("حجم التذكرة:", [6, 7, 8, 9, 10], horizontal=True)
    
    with c_check2:
        chk_numbers = st.multiselect(
            f"اختر {chk_size} أرقام:",
            options=list(range(1, 33)),
            max_selections=chk_size,
            key="checker_multiselect"
        )
    
    if st.button("🔎 فحص شامل", type="primary", use_container_width=True):
        if len(chk_numbers) != chk_size:
            st.error(f"⚠️ يجب اختيار {chk_size} أرقام")
        else:
            sorted_chk = sorted(chk_numbers)
            st.success(f"✅ فحص: {sorted_chk}")
            
            st.markdown("### 1️⃣ التطابقات")
            matches = analyzer.check_matches_history(sorted_chk)
            found_any = False
            
            for count in [6, 5, 4, 3]:
                res_list = matches[count]
                if res_list:
                    found_any = True
                    with st.expander(f"🌟 {count} أرقام ({len(res_list)} مرة)", expanded=True):
                        for item in res_list:
                            st.markdown(f"- سحب {item['draw_id']}: {item['matched_nums']}")
            
            if not found_any:
                st.info("✅ تذكرة نظيفة!")
            
            st.divider()
            
            st.markdown("### 2️⃣ التكرار")
            freq_df = analyzer.get_numbers_frequency_stats(sorted_chk)
            col_f1, col_f2 = st.columns([1, 2])
            
            with col_f1:
                st.dataframe(freq_df, hide_index=True, use_container_width=True)
            
            with col_f2:
                fig = px.bar(freq_df, x='الرقم', y='عدد مرات الظهور', color='التصنيف',
                            color_discrete_map={'ساخن': '#ef4444', 'بارد': '#3b82f6'})
                st.plotly_chart(fig, use_container_width=True)
            
            st.divider()
            
            st.markdown("### 3️⃣ المتتاليات")
            seq_results = analyzer.analyze_sequences_history(sorted_chk)
            
            if not seq_results:
                st.write("لا توجد متتاليات")
            else:
                for seq_tuple, data in seq_results.items():
                    st.markdown(f"#### {seq_tuple}")
                    st.write(f"- ظهرت: {data['full_count']} مرة")
                    if data['full_draws']:
                        st.caption(f"السحوبات: {data['full_draws']}")


def render_deep_analytics(analyzer: AdvancedAnalyzer):
    st.header("📈 التحليلات العميقة")
    
    analysis_tabs = st.tabs([
        "📅 زمني",
        "🔗 ارتباط",
        "⏳ فجوات",
        "📊 انحراف"
    ])
    
    with analysis_tabs[0]:
        st.subheader("📅 التحليل الزمني")
        
        period_type = st.radio("حسب:", ["شهر", "سنة"], horizontal=True)
        period_map = {"شهر": "month", "سنة": "year"}
        
        temporal_df = analyzer.get_temporal_analysis(period_map[period_type])
        
        st.markdown("### الأرقام الأكثر ظهوراً")
        
        pivot_table = temporal_df.pivot_table(
            index='number',
            columns='period',
            values='frequency',
            fill_value=0
        )
        
        fig_temporal = go.Figure()
        
        for period in pivot_table.columns:
            fig_temporal.add_trace(go.Bar(
                name=str(period),
                x=pivot_table.index.tolist(),
                y=pivot_table[period].tolist()
            ))
        
        fig_temporal.update_layout(
            barmode='group',
            xaxis_title='الرقم',
            yaxis_title='التكرار',
            height=400
        )
        
        st.plotly_chart(fig_temporal, use_container_width=True)
    
    with analysis_tabs[1]:
        st.subheader("🔗 مصفوفة الارتباط")
        
        corr_type = st.radio("النوع:", ["أزواج", "ثلاثيات"], horizontal=True)
        
        if corr_type == "أزواج":
            st.markdown("### أكثر 20 زوج")
            top_pairs = analyzer.get_top_pairs(20)
            
            df_pairs = pd.DataFrame([
                {
                    'الزوج': f"{pair[0]}-{pair[1]}",
                    'التكرار': count
                }
                for pair, count in top_pairs
            ])
            
            col_p1, col_p2 = st.columns([1, 1])
            
            with col_p1:
                st.dataframe(df_pairs, hide_index=True, use_container_width=True)
            
            with col_p2:
                fig_pairs = px.bar(df_pairs, x='الزوج', y='التكرار', 
                                  color='التكرار', color_continuous_scale='Viridis')
                st.plotly_chart(fig_pairs, use_container_width=True)
        
        else:
            st.markdown("### أكثر 10 ثلاثيات")
            top_triplets = analyzer.get_top_triplets(10)
            
            df_triplets = pd.DataFrame([
                {
                    'الثلاثية': f"{t[0]}-{t[1]}-{t[2]}",
                    'التكرار': count
                }
                for t, count in top_triplets
            ])
            
            st.dataframe(df_triplets, hide_index=True, use_container_width=True)
    
    with analysis_tabs[2]:
        st.subheader("⏳ تحليل الفجوات")
        
        gaps_df = pd.DataFrame([
            {'الرقم': num, 'الفجوة': gap}
            for num, gap in sorted(analyzer.gaps.items(), key=lambda x: x[1], reverse=True)
        ])
        
        alert_threshold = st.slider("حد التنبيه:", 5, 50, 15)
        delayed_numbers = gaps_df[gaps_df['الفجوة'] >= alert_threshold]
        
        if not delayed_numbers.empty:
            st.warning(f"⚠️ {len(delayed_numbers)} رقم متأخر!")
            st.dataframe(delayed_numbers, hide_index=True, use_container_width=True)
        
        st.divider()
        
        fig_gaps = px.bar(gaps_df, x='الرقم', y='الفجوة', 
                         color='الفجوة', color_continuous_scale='Reds')
        fig_gaps.add_hline(y=alert_threshold, line_dash="dash", 
                          line_color="red")
        st.plotly_chart(fig_gaps, use_container_width=True)
    
    with analysis_tabs[3]:
        st.subheader("📊 الانحراف المعياري")
        
        std_data = analyzer.calculate_standard_deviation()
        
        std_df = pd.DataFrame([
            {'الرقم': num, 'الانحراف': std}
            for num, std in sorted(std_data['number_std'].items())
        ])
        
        fig_std = px.line(std_df, x='الرقم', y='الانحراف', markers=True)
        st.plotly_chart(fig_std, use_container_width=True)
        
        col_s1, col_s2 = st.columns(2)
        with col_s1:
            st.metric("متوسط المجموع", f"{std_data['sum_mean']:.2f}")
        with col_s2:
            st.metric("الانحراف", f"{std_data['sum_std']:.2f}")


def render_probability_simulator(analyzer: AdvancedAnalyzer):
    st.header("🧮 الاحتمالات والمحاكي")
    
    prob_tabs = st.tabs([
        "📊 الاحتمالات",
        "🎲 المحاكي",
        "💰 القيمة المتوقعة"
    ])
    
    with prob_tabs[0]:
        st.subheader("📊 حساب الاحتمالية")
        
        ticket_size_prob = st.slider("حجم التذكرة:", 6, 10, 6)
        
        probs_data = []
        for match in [3, 4, 5, 6]:
            prob = ProbabilityCalculator.calculate_match_probability(ticket_size_prob, match)
            prob_percent = prob * 100
            one_in = 1 / prob if prob > 0 else float('inf')
            
            probs_data.append({
                'التطابق': f'{match} أرقام',
                'الاحتمالية': f'{prob_percent:.6f}%',
                'واحد في': f'1 في {one_in:,.0f}' if one_in != float('inf') else 'مستحيل'
            })
        
        df_probs = pd.DataFrame(probs_data)
        st.dataframe(df_probs, hide_index=True, use_container_width=True)
    
    with prob_tabs[1]:
        st.subheader("🎲 محاكي السحوبات")
        
        col_sim1, col_sim2 = st.columns([1, 1])
        
        with col_sim1:
            sim_ticket = st.multiselect(
                "اختر 6 أرقام:",
                options=list(range(1, 33)),
                max_selections=6,
                key="simulator_multiselect"
            )
            
            num_sims = st.selectbox(
                "عدد المحاكاة:",
                [1000, 10000, 50000, 100000],
                index=1
            )
            
            if st.button("🚀 محاكاة", type="primary") and len(sim_ticket) == 6:
                simulator = DrawSimulator(analyzer)
                
                progress_bar = st.progress(0)
                progress_text = st.empty()
                
                def sim_progress(current, total):
                    progress_bar.progress(current / total)
                    progress_text.text(f"{current:,} / {total:,}")
                
                with st.spinner("جاري..."):
                    results = simulator.simulate_draws(num_sims, sorted(sim_ticket), sim_progress)
                
                progress_bar.empty()
                progress_text.empty()
                
                st.success(f"✅ {num_sims:,} محاكاة!")
                st.session_state['sim_results'] = results
        
        with col_sim2:
            if 'sim_results' in st.session_state:
                results = st.session_state['sim_results']
                
                st.markdown("### 📊 النتائج:")
                
                results_data = []
                for match_level in [3, 4, 5, 6]:
                    data = results['results'][match_level]
                    results_data.append({
                        'التطابق': f'{match_level} أرقام',
                        'العدد': data['count'],
                        'النسبة': f"{data['percentage']:.4f}%",
                        'النظرية': f"{data['theoretical']:.4f}%"
                    })
                
                df_results = pd.DataFrame(results_data)
                st.dataframe(df_results, hide_index=True, use_container_width=True)
    
    with prob_tabs[2]:
        st.subheader("💰 القيمة المتوقعة")
        
        col_ev1, col_ev2 = st.columns([1, 1])
        
        with col_ev1:
            ticket_size_ev = st.slider("حجم التذكرة:", 6, 10, 6, key='ev_size')
            jackpot_amount = st.number_input(
                "الجائزة الكبرى:",
                min_value=100000,
                max_value=10000000,
                value=1000000,
                step=100000
            )
            
            ev_data = ProbabilityCalculator.calculate_expected_value(ticket_size_ev, jackpot_amount)
            
            st.metric("التكلفة", f"{ev_data['ticket_cost']} د")
            st.metric("القيمة المتوقعة", f"{ev_data['expected_value']:.2f} د")
            st.metric("الصافي", f"{ev_data['net_expected_value']:.2f} د")
            
            roi_color = "green" if ev_data['roi'] > 0 else "red"
            st.markdown(f"**العائد:** :{roi_color}[{ev_data['roi']:.2f}%]")
        
        with col_ev2:
            st.markdown("### 🎁 التفصيل:")
            
            breakdown_data = []
            for level, data in ev_data['breakdown'].items():
                prize_display = f"{data['prize']:,}" if isinstance(data['prize'], (int, float)) else data['prize']
                breakdown_data.append({
                    'المستوى': level,
                    'الجائزة': prize_display,
                    'المساهمة': f"{data['contribution']:.4f} د"
                })
            
            df_breakdown = pd.DataFrame(breakdown_data)
            st.dataframe(df_breakdown, hide_index=True, use_container_width=True)


def render_user_guide():
    st.header("📖 دليل الاستخدام")
    
    with st.expander("📊 **Dashboard**", expanded=True):
        st.markdown("""
        ### الوظيفة:
        نظرة سريعة على البيانات
        
        ### المحتوى:
        - إحصائيات سريعة
        - أكثر/أقل الأرقام ظهوراً
        - خريطة حرارية
        - آخر 5 سحوبات
        """)
    
    with st.expander("🎰 **المولد الذكي**"):
        st.markdown("""
        ### الإعدادات:
        - **الاستراتيجية:** ساخنة/باردة/متوازنة
        - **المتوسط:** الالتزام بمتوسط محدد
        - **حجم التذكرة:** 6-10
        - **المتتاليات والظلال**
        - **تثبيت أرقام**
        
        ### الخطوات:
        1. اختر المعايير
        2. فحص الجدوى
        3. توليد
        4. تصدير
        """)
    
    with st.expander("🔍 **الفاحص**"):
        st.markdown("""
        ### الفحوصات:
        - التطابقات التاريخية
        - تكرار الأرقام
        - المتتاليات
        """)
    
    with st.expander("📈 **التحليلات**"):
        st.markdown("""
        - **زمني:** توزيع حسب الشهر/السنة
        - **ارتباط:** أزواج وثلاثيات شائعة
        - **فجوات:** آخر ظهور
        - **انحراف:** قياس التشتت
        """)
    
    with st.expander("🧮 **الاحتمالات**"):
        st.markdown("""
        - حساب احتمالية التطابق
        - محاكاة السحوبات
        - القيمة المتوقعة
        """)
    
    st.divider()
    
    st.warning("""
    ⚠️ **إخلاء مسؤولية:**
    
    هذا تطبيق تحليلي تعليمي. اليانصيب لعبة حظ عشوائية.
    لا توجد خوارزمية تضمن الفوز. العب بمسؤولية!
    """)


# ==============================================================================
# 9. الدالة الرئيسية
# ==============================================================================

def main():
    st.set_page_config(
        page_title="🎰 اليانصيب الأردني",
        page_icon="🎲",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    initialize_session_state()
    apply_theme()
    
    # Sidebar
    with st.sidebar:
        st.title("⚙️ الإعدادات")
        
        theme_label = "🌙 داكن" if st.session_state.theme == 'light' else "☀️ فاتح"
        if st.button(theme_label, use_container_width=True):
            st.session_state.theme = 'dark' if st.session_state.theme == 'light' else 'light'
            st.rerun()
        
        st.divider()
        
        st.subheader("🎨 الألوان")
        st.session_state.hot_color = st.color_picker("ساخنة", st.session_state.hot_color)
        st.session_state.cold_color = st.color_picker("باردة", st.session_state.cold_color)
        st.session_state.balanced_color = st.color_picker("متوازنة", st.session_state.balanced_color)
        
        st.divider()
        
        st.subheader("📂 البيانات")
        uploaded_file = st.file_uploader("رفع ملف", type=['xlsx', 'xls', 'csv'])
        
        if uploaded_file:
            with st.spinner("جاري التحميل..."):
                df, msg = load_and_process_data(uploaded_file)
                if df is not None:
                    st.session_state.history_df = df
                    st.session_state.analyzer = AdvancedAnalyzer(df)
                    st.session_state.generator = SmartGenerator(st.session_state.analyzer)
                    st.success(f"✅ {len(df)} سحب")
                else:
                    st.error(msg)
    
    if st.session_state.history_df is None:
        st.warning("⚠️ يرجى رفع ملف البيانات")
        st.info("""
        ### 📋 المتطلبات:
        - Excel (.xlsx) أو CSV
        - الأعمدة: N1, N2, N3, N4, N5, N6
        - الأرقام: 1-32
        """)
        return
    
    analyzer = st.session_state.analyzer
    generator = st.session_state.generator
    
    st.title("🎰 اليانصيب الأردني المتطور")
    st.markdown("**النسخة 2.0** - تحليلات متقدمة")
    
    tabs = st.tabs([
        "📊 Dashboard",
        "🎰 المولد",
        "🔍 الفاحص",
        "📈 التحليلات",
        "🧮 الاحتمالات",
        "📖 الدليل"
    ])
    
    with tabs[0]:
        render_dashboard(analyzer)
    
    with tabs[1]:
        render_smart_generator(analyzer, generator)
    
    with tabs[2]:
        render_checker(analyzer)
    
    with tabs[3]:
        render_deep_analytics(analyzer)
    
    with tabs[4]:
        render_probability_simulator(analyzer)
    
    with tabs[5]:
        render_user_guide()
    
    st.markdown("""
    <div class="footer">
        <b>برمجة: محمد العمري</b> | النسخة 2.0.0 | فبراير 2026
    </div>
    """, unsafe_allow_html=True)


if __name__ == "__main__":
    main()
